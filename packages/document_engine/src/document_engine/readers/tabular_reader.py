import csv
import logging
from pathlib import Path

import openpyxl
from tabulate import tabulate

from ..schema import Fragment

logger = logging.getLogger(__name__)


class TabularReader:
    """Reader：负责将 Excel 和 CSV 转换为 Fragment 列表"""

    def parse(self, file_path: str | Path) -> list[Fragment]:
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".csv":
            return self._parse_csv(path)
        elif suffix in [".xlsx", ".xlsm"]:
            return self._parse_excel(path)
        else:
            logger.warning(f"Unsupported file type: {suffix}")
            return []

    def _parse_csv(self, path: Path) -> list[Fragment]:
        fragments = []
        try:
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                data = list(reader)
                if not data:
                    return []

                md_table = tabulate(data, headers="firstrow", tablefmt="pipe")
                fragments.append(
                    Fragment(
                        type="table",
                        y0=0.0,
                        content=md_table,
                        meta={"source_file": path.name, "sheet_name": "default"},
                    )
                )
        except Exception as e:
            logger.error(f"Error parsing CSV {path}: {e}")

        return fragments

    def _parse_excel(self, path: Path) -> list[Fragment]:
        fragments = []
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    # 过滤掉全空的行
                    if any(cell is not None for cell in row):
                        data.append(list(row))

                if not data:
                    continue

                md_table = tabulate(data, headers="firstrow", tablefmt="pipe")
                fragments.append(
                    Fragment(
                        type="table",
                        y0=0.0,
                        content=md_table,
                        meta={"source_file": path.name, "sheet_name": sheet_name},
                    )
                )
        except Exception as e:
            logger.error(f"Error parsing Excel {path}: {e}")

        return fragments
